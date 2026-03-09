# inventory_manager.py  (New System)
# Reads the processed meal DataFrames (from auto_svaadh_summary) to calculate
# packaging unit requirements and costs.
#
# New column names used:
#   Chapati, Without_Oil_Chapati, Phulka, Ghee_Phulka, Jowar_Bhakri, Bajra_Bhakri
#   Dry_Sabji_Mini, Dry_Sabji_Full, Curry_Sabji_Mini, Curry_Sabji_Full
#   Dal, Rice, Salad, Curd
#   BF_Qty_1 … BF_Qty_4  (breakfast quantities)

import os
import glob
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import warnings

warnings.filterwarnings("ignore", category=UserWarning)

# ── CONFIG ───────────────────────────────────────────────────
LOCAL_ROOT     = os.path.join(os.getcwd(), "Processed_Orders")
OUTPUT_FOLDER  = os.path.join(LOCAL_ROOT, "Inventory_Summary")

UNIT_COSTS = {
    "Aluminium pouch":          0.55,
    "100ml plastic container":  1.77,
    "250ml plastic container":  3.12,
    "250ml aluminium container":1.50,
    "400ml aluminium container":2.00,
    "50ml plastic":             1.10,
}

# Each tuple: (list_of_col_names, packaging_type, rule)
# rule = "one_per_row" → 1 unit per non-zero cell (rotis/salad)
#        "sum_qty"     → sum the quantity column
MAPPINGS = [
    (
        ["Chapati","Without_Oil_Chapati","Phulka","Ghee_Phulka",
         "Jowar_Bhakri","Bajra_Bhakri","Salad"],
        "Aluminium pouch", "one_per_row"
    ),
    (
        ["Dry_Sabji_Mini","Curry_Sabji_Mini"],
        "100ml plastic container", "sum_qty"
    ),
    (
        ["Dry_Sabji_Full","Curry_Sabji_Full","Dal"],
        "250ml plastic container", "sum_qty"
    ),
    (
        ["Rice"],
        "250ml aluminium container", "sum_qty"
    ),
    (
        ["Curd"],
        "50ml plastic", "sum_qty"
    ),
    # Breakfast items — each BF_Qty_N column counts as individual portions
    (
        ["BF_Qty_1","BF_Qty_2","BF_Qty_3","BF_Qty_4"],
        "Aluminium pouch", "sum_qty"
    ),
]


# ── CORE COUNTING ────────────────────────────────────────────

def _safe_series(df, col):
    """Return series for col, excluding the last 2 summary rows (totals)."""
    if col not in df.columns:
        return pd.Series(dtype=float)
    s = df[col]
    return s.iloc[:-2] if len(s) > 2 else s


def _count_one_per_row(series: pd.Series) -> int:
    """Count rows where qty > 0 (one packaging unit per non-zero cell)."""
    numeric = pd.to_numeric(series, errors="coerce").fillna(0)
    return int((numeric > 0).sum())


def _sum_qty(series: pd.Series) -> float:
    """Sum all quantity values."""
    return round(float(pd.to_numeric(series, errors="coerce").fillna(0).sum()), 2)


def apply_mappings(all_dfs: dict) -> dict:
    """
    all_dfs: {"Breakfast": df, "Lunch": df, "Dinner": df}
    Returns {packaging_item: total_units_used}
    """
    counts = {k: 0.0 for k in UNIT_COSTS}

    for meal_name, df in all_dfs.items():
        if df is None or df.empty:
            continue

        for col_names, inv_key, rule in MAPPINGS:
            for col in col_names:
                series = _safe_series(df, col)
                if series.empty:
                    continue
                if rule == "one_per_row":
                    counts[inv_key] += _count_one_per_row(series)
                else:
                    counts[inv_key] += _sum_qty(series)

    return {k: (int(v) if v % 1 == 0 else round(v, 2)) for k, v in counts.items()}


# ── OUTPUT ───────────────────────────────────────────────────

def write_inventory_summary(date_str: str, all_dfs: dict):
    """
    date_str: "YYYY-MM-DD"
    all_dfs:  {"Breakfast": df, "Lunch": df, "Dinner": df}
              Each df is the filtered meal DataFrame from process_data().
    """
    counts     = apply_mappings(all_dfs)
    rows       = []
    grand_total = 0.0

    for item, cost in UNIT_COSTS.items():
        used  = counts.get(item, 0)
        total = round(used * cost, 2)
        grand_total += total
        rows.append({
            "Item":            item,
            "Units Used":      used,
            "Unit Cost (₹)":   cost,
            "Total Cost (₹)":  total,
        })

    rows.append({
        "Item":           "Grand Total",
        "Units Used":     "",
        "Unit Cost (₹)":  "",
        "Total Cost (₹)": round(grand_total, 2),
    })

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    out_file = os.path.join(OUTPUT_FOLDER, f"inventory_summary_{date_str}.xlsx")
    pd.DataFrame(rows).to_excel(out_file, index=False)
    _format_excel(out_file)

    _update_monthly_yearly(date_str, grand_total)
    print(f"✅ Inventory summary → {out_file}")


def _update_monthly_yearly(date_str: str, total_cost: float):
    dt         = datetime.strptime(date_str, "%Y-%m-%d")
    year       = dt.year
    month_name = dt.strftime("%B")
    month_id   = dt.strftime("%Y-%m")

    # Monthly
    month_dir  = os.path.join(LOCAL_ROOT, str(year), month_name)
    os.makedirs(month_dir, exist_ok=True)
    month_file = os.path.join(month_dir, f"inventory_monthly_summary_{month_id}.xlsx")
    new_day    = pd.DataFrame([{"Date": date_str, "Total Cost (₹)": round(total_cost, 2)}])

    if os.path.exists(month_file):
        df_m = pd.read_excel(month_file)
        df_m = df_m[df_m["Date"] != "Grand Total"]
        df_m = pd.concat([df_m, new_day]).drop_duplicates(subset=["Date"], keep="last")
    else:
        df_m = new_day

    m_total = round(df_m["Total Cost (₹)"].sum(), 2)
    df_m    = pd.concat([df_m, pd.DataFrame([{"Date":"Grand Total","Total Cost (₹)":m_total}])])
    df_m.to_excel(month_file, index=False)
    _format_excel(month_file)

    # Yearly
    year_file  = os.path.join(LOCAL_ROOT, str(year), f"inventory_yearly_summary_{year}.xlsx")
    new_month  = pd.DataFrame([{"Month": month_name, "Total Cost (₹)": m_total}])

    if os.path.exists(year_file):
        df_y = pd.read_excel(year_file)
        df_y = df_y[df_y["Month"] != "Grand Total"]
        df_y = df_y[df_y["Month"] != month_name]
        df_y = pd.concat([df_y, new_month])
    else:
        df_y = new_month

    y_total = round(df_y["Total Cost (₹)"].sum(), 2)
    df_y    = pd.concat([df_y, pd.DataFrame([{"Month":"Grand Total","Total Cost (₹)":y_total}])])
    df_y.to_excel(year_file, index=False)
    _format_excel(year_file)


def _format_excel(path: str):
    wb   = load_workbook(path)
    ws   = wb.active
    thin = Side(border_style="thin", color="000000")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Bold the header row (row 1) and any Grand Total row — never assume max_row
        first_cell_val = str(row[0].value).strip() if row[0].value is not None else ""
        if row[0].row == 1 or first_cell_val in ("Grand Total", "── TOTAL ──"):
            for cell in row:
                cell.font = Font(bold=True)
    wb.save(path)


if __name__ == "__main__":
    # Standalone test — point to any saved customer_summary xlsx
    print("Run via auto_svaadh_summary.ipynb for full pipeline.")
