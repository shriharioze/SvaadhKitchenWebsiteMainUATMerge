# customer_info_manager.py  (New System)
# Reads SK_Orders tab, builds customer_info_master.xlsx from first-time customers
# Column reference: Phone, Customer_Name, Full_Address, Maps_Link, Landmark,
#                   Payment_Freq, First_Time, Area, Wing, Flat, Floor, Society

import os
import gspread
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from google.oauth2.service_account import Credentials

# ── CONFIG ───────────────────────────────────────────────────
SERVICE_ACCOUNT_FILE = "service_account.json"
SHEET_ID             = "17X7JOrMe1Oj_QykH7mk6UGoBjuguLfyC1RmKYATDXlI"   # ← new sheet
TAB_ORDERS           = "SK_Orders"
CUSTOMER_INFO_DIR    = os.path.join(os.getcwd(), "Processed_Orders", "Customer_Info")
CUSTOMER_INFO_FILE   = os.path.join(CUSTOMER_INFO_DIR, "customer_info_master.xlsx")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

def _auth():
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return gspread.authorize(creds)

def update_customer_info():
    os.makedirs(CUSTOMER_INFO_DIR, exist_ok=True)
    print("📋 Updating Customer Info...")

    gc  = _auth()
    sh  = gc.open_by_key(SHEET_ID)
    ws  = sh.worksheet(TAB_ORDERS)
    data = ws.get_all_records()
    df  = pd.DataFrame(data)

    if df.empty:
        print("ℹ️  No orders in sheet yet.")
        return

    df.columns = df.columns.str.strip()

    # Only first-time customers
    if "First_Time" not in df.columns:
        print("⚠️  First_Time column missing.")
        return

    new_df = df[df["First_Time"].astype(str).str.lower() == "yes"].copy()

    if new_df.empty:
        print("ℹ️  No new first-time customers.")
        return

    # Map payment freq → Daily? flag (keeps backward compat with dispatch/accounting)
    freq_map = {
        "Daily Payment":                            "Yes",
        "Prepaid Wallet":                           "No",
        "Will decide later?":                       "Yes",
    }
    new_df["Daily?"] = new_df["Payment_Freq"].map(freq_map).fillna("Yes")

    # Standardise name capitalisation
    new_df["Customer_Name"] = (
        new_df["Customer_Name"].astype(str).str.strip().str.title()
    )

    # Keep only the columns we care about
    keep = [
        "Customer_Name", "Phone", "Full_Address",
        "Area", "Wing", "Flat", "Floor", "Society",
        "Maps_Link", "Landmark", "Payment_Freq", "Daily?"
    ]
    new_df = new_df[[c for c in keep if c in new_df.columns]].copy()

    # Merge with existing master (phone as unique key, keep first occurrence = oldest)
    if os.path.exists(CUSTOMER_INFO_FILE):
        existing = pd.read_excel(CUSTOMER_INFO_FILE)
        if "Daily?" not in existing.columns:
            existing["Daily?"] = "Yes"
        existing["Customer_Name"] = (
            existing["Customer_Name"].astype(str).str.strip().str.title()
        )
        merged = (
            pd.concat([existing, new_df])
            .drop_duplicates(subset=["Phone"], keep="first")
            .reset_index(drop=True)
        )
    else:
        merged = new_df

    merged.to_excel(CUSTOMER_INFO_FILE, index=False)
    _format(CUSTOMER_INFO_FILE)
    print(f"✅ Customer master updated → {len(merged)} total customers")

def _format(path):
    wb  = load_workbook(path)
    ws  = wb.active
    thin = Side(border_style="thin", color="000000")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(path)

if __name__ == "__main__":
    update_customer_info()
