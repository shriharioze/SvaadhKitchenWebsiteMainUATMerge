# auto_svaadh_summary.py
# ============================================================
# SVAADH KITCHEN — Master Orchestrator (New System)
# Reads SK_Orders from Google Sheet, runs all sub-modules.
# Save as auto_svaadh_summary.ipynb in Colab for interactive use.
# ============================================================

# ── CELL 1: Mount Drive & set working directory ──────────────
import os
from google.colab import drive

if not os.path.exists("/content/drive"):
    drive.mount("/content/drive")

COLAB_DRIVE_ROOT = "/content/drive/MyDrive/Svaadh Kitchen/New System"
os.chdir(COLAB_DRIVE_ROOT)
print(f"📁 Working directory: {os.getcwd()}")


# ── CELL 2: Imports & config ──────────────────────────────────
import time
import importlib.util
import pandas as pd
import gspread
from datetime import datetime, timedelta, date
from google.oauth2.service_account import Credentials

SERVICE_ACCOUNT_FILE = "service_account.json"
SHEET_ID             = "17X7JOrMe1Oj_QykH7mk6UGoBjuguLfyC1RmKYATDXlI"   # ← paste your new Sheet ID
TAB_ORDERS           = "SK_Orders"
INTERVAL             = 300   # seconds between refreshes (5 min)

LOCAL_ROOT = os.path.join(os.getcwd(), "Processed_Orders")
os.makedirs(LOCAL_ROOT, exist_ok=True)

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/drive.file",
]
creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
gc    = gspread.authorize(creds)
print("✅ Auth OK")


# ── CELL 3: Load sub-modules ──────────────────────────────────
def load_module(name, path):
    spec   = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module

customer_info_mod = load_module("customer_info_manager", "customer_info_manager.py")
kitchen_live_mod  = load_module("kitchen_live_manager",  "kitchen_live_manager.py")
inventory_mod     = load_module("inventory_manager",     "inventory_manager.py")
dispatch_mod      = load_module("dispatch_manager",      "dispatch_manager.py")
label_gen_mod     = load_module("label_generator",       "label_generator.py")
print("✅ All modules loaded")


# ── CELL 4: Ask processing date ───────────────────────────────
input_date = input("📅 Enter date to process (YYYY-MM-DD) or press Enter for today: ").strip()
if input_date:
    process_date = datetime.strptime(input_date, "%Y-%m-%d").date()
else:
    process_date = date.today()

date_str     = process_date.strftime("%Y-%m-%d")
year_folder  = str(process_date.year)
month_folder = process_date.strftime("%B")
print(f"🔎 Processing: {date_str}")


# ── CELL 5: Core processing functions ────────────────────────
def fetch_sheet() -> pd.DataFrame:
    sh  = gc.open_by_key(SHEET_ID)
    ws  = sh.worksheet(TAB_ORDERS)
    df  = pd.DataFrame(ws.get_all_records())
    if df.empty:
        return df
    df.columns = df.columns.str.strip()
    df["Order_Date"] = pd.to_datetime(df["Order_Date"], errors="coerce").dt.date
    return df


def process_data(df: pd.DataFrame) -> dict:
    """
    Filter for process_date, split into meal DataFrames,
    compute per-customer daily totals.

    Returns dict with keys:
      "Breakfast", "Lunch", "Dinner"         → raw meal DataFrames (all rows)
      "Summary_Breakfast", ...               → customer order summary
      "Daily_Customer_Total"                 → per-customer totals across all meals
    """
    results = {}

    day_df = df[df["Order_Date"] == process_date].copy()
    if day_df.empty:
        print(f"⚠️  No rows found for {date_str}")
        return results

    # Standardise name
    day_df["Customer_Name"] = (
        day_df["Customer_Name"].astype(str).str.strip().str.title()
    )

    # Numeric item columns
    item_cols = [
        "Chapati","Without_Oil_Chapati","Phulka","Ghee_Phulka",
        "Jowar_Bhakri","Bajra_Bhakri",
        "Dry_Sabji_Mini","Dry_Sabji_Full",
        "Curry_Sabji_Mini","Curry_Sabji_Full",
        "Dal","Rice","Salad","Curd",
        "BF_Qty_1","BF_Qty_2","BF_Qty_3","BF_Qty_4",
    ]
    for col in item_cols:
        if col in day_df.columns:
            day_df[col] = pd.to_numeric(day_df[col], errors="coerce").fillna(0)

    # Split by meal
    for meal in ["Breakfast", "Lunch", "Dinner"]:
        meal_df = day_df[
            day_df["Meal_Type"].astype(str).str.strip().str.title() == meal
        ].copy()

        if meal_df.empty:
            continue

        # Add totals row
        totals      = meal_df[item_cols].select_dtypes(include="number").sum()
        total_row   = pd.Series({c: "" for c in meal_df.columns})
        total_row.update(totals)
        total_row["Customer_Name"] = "TOTAL"
        meal_df = pd.concat(
            [meal_df, pd.DataFrame([total_row])], ignore_index=True
        )

        results[meal] = meal_df

        # Customer order summaries (name + item summary text)
        summary_rows = []
        for _, row in meal_df.iterrows():
            name = str(row.get("Customer_Name","")).strip()
            if name in ("", "TOTAL", "nan"):
                continue
            try:
                import json
                items_obj = json.loads(str(row.get("Items_JSON", "{}")))
                summary   = ", ".join(
                    f"{int(v)}×{k}" for k,v in items_obj.items() if float(v) > 0
                )
            except Exception:
                summary = "—"
            summary_rows.append({
                "Meal Type":        meal,
                "Customer Name":    name,
                "Order Summary":    summary,
                "Special Notes":    str(row.get("Special_Notes","")).strip(),
                "Full Address":     str(row.get("Full_Address","")).strip(),
                "Phone":            str(row.get("Phone","")).strip(),
            })
        results[f"Summary_{meal}"] = pd.DataFrame(summary_rows)

    # Daily customer totals (across all meals)
    day_df["Net_Total"] = pd.to_numeric(day_df["Net_Total"], errors="coerce").fillna(0)
    daily_total = (
        day_df.groupby(["Customer_Name","Phone","Payment_Freq"], as_index=False)["Net_Total"]
        .sum()
        .rename(columns={"Net_Total": "Total (All Meals)"})
    )
    daily_total["Payment_Status"] = "Pending"
    total_sum = daily_total["Total (All Meals)"].sum()
    daily_total = pd.concat([
        daily_total,
        pd.DataFrame([{
            "Customer_Name":  "── TOTAL ──",
            "Phone":          "",
            "Payment_Freq":   "",
            "Total (All Meals)": total_sum,
            "Payment_Status": "",
        }])
    ], ignore_index=True)
    results["Daily_Customer_Total"] = daily_total

    return results


def save_excel(results: dict):
    """Save all result DataFrames to customer_summary_{date}.xlsx
    Writes to a .tmp file first, then renames — so the final file is never
    left in a half-written state if Excel has it open or a crash occurs.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Border, Side

    local_dir  = os.path.join(LOCAL_ROOT, year_folder, month_folder)
    os.makedirs(local_dir, exist_ok=True)
    out_path   = os.path.join(local_dir, f"customer_summary_{date_str}.xlsx")
    tmp_path   = out_path + ".tmp"

    # Write to temp file (5 retries if the temp itself is locked)
    for attempt in range(5):
        try:
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                for name, df_data in results.items():
                    if isinstance(df_data, pd.DataFrame) and not df_data.empty:
                        df_data.to_excel(writer, index=False, sheet_name=str(name)[:31])
            break
        except PermissionError:
            print(f"⚠️  Temp file locked (attempt {attempt+1}/5). Retrying…")
            time.sleep(5)
    else:
        print("❌ Could not write temp file after 5 attempts. Skipping save.")
        return None

    # Format the temp file
    try:
        wb   = load_workbook(tmp_path)
        thin = Side(border_style="thin", color="000000")
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = max_len + 2
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        wb.save(tmp_path)
    except Exception as e:
        print(f"⚠️  Formatting error: {e}")

    # Atomic rename: replace the final file only after the temp is fully ready
    try:
        if os.path.exists(out_path):
            os.remove(out_path)
        os.rename(tmp_path, out_path)
    except Exception as e:
        print(f"⚠️  Rename failed: {e}. Temp file kept at {tmp_path}")
        return tmp_path

    print(f"✅ Excel summary saved → {out_path}")
    return out_path


# ── CELL 6: Main loop ─────────────────────────────────────────
def main_loop():
    print(f"\n🚀 Starting live refresh loop for {date_str}")
    print(f"   Refresh interval: {INTERVAL//60} min  |  Ctrl+C to stop\n")

    while True:
        try:
            ist = datetime.utcnow() + timedelta(hours=5, minutes=30)
            print(f"\n🔁 Refresh at {ist.strftime('%Y-%m-%d %H:%M:%S')} IST")

            # 1. Fetch sheet
            full_df = fetch_sheet()
            if full_df.empty:
                print("ℹ️  Sheet is empty.")
                time.sleep(INTERVAL)
                continue

            # 2. Filter to today
            day_df  = full_df[full_df["Order_Date"] == process_date].copy()
            if day_df.empty:
                print(f"ℹ️  No orders for {date_str}.")
                time.sleep(INTERVAL)
                continue

            print(f"   📊 {len(day_df)} rows for {date_str}")

            # 3. Process
            results = process_data(full_df)
            if not results:
                time.sleep(INTERVAL)
                continue

            # 4. Save Excel summary
            save_excel(results)

            # 5. Customer info master
            customer_info_mod.update_customer_info()

            # 6. Inventory summary — pass meal DataFrames directly
            all_dfs = {
                "Breakfast": results.get("Breakfast"),
                "Lunch":     results.get("Lunch"),
                "Dinner":    results.get("Dinner"),
            }
            inventory_mod.write_inventory_summary(date_str, all_dfs)

            # 7. Kitchen live dashboard
            kitchen_live_mod.update_live_dashboard(results)

            # 8. Dispatch file — pass filtered day_df directly
            dispatch_mod.generate_dispatch_file(date_str, day_df)

            print(f"✅ All modules refreshed.")
            print(f"🕒 Next refresh in {INTERVAL//60} min…\n")
            time.sleep(INTERVAL)

        except KeyboardInterrupt:
            print("\n⛔ Stopped by user.")
            break
        except Exception as e:
            print(f"❌ Error: {e}")
            time.sleep(60)


# ── CELL 7: RUN ───────────────────────────────────────────────
if __name__ == "__main__":
    main_loop()
