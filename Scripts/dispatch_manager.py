# dispatch_manager.py  (New System)
# Reads SK_Orders directly — address is already in every row.
# No separate customer_info_master.xlsx lookup needed for addresses.
# Produces driver_dispatch_summary.xlsx with one sheet per meal.

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from datetime import datetime

# ── CONFIG ───────────────────────────────────────────────────
# Use the script's own directory so the output path is stable
# regardless of which directory the caller imports from.
_SCRIPT_DIR       = os.path.dirname(os.path.abspath(__file__))
LOCAL_ROOT        = os.path.join(_SCRIPT_DIR, "Processed_Orders")
DISPATCH_LINK_DIR = os.path.join(LOCAL_ROOT, "Dispatch_Link")
DISPATCH_LINK_FILE= os.path.join(DISPATCH_LINK_DIR, "driver_dispatch_summary.xlsx")


def generate_dispatch_file(date_str: str, orders_df: pd.DataFrame):
    """
    orders_df: the full SK_Orders DataFrame already filtered for date_str
               (passed in from auto_svaadh_summary — no extra sheet read needed).
    """
    if orders_df is None or orders_df.empty:
        print(f"⚠️  No orders found for {date_str}. Dispatch file not generated.")
        return

    os.makedirs(DISPATCH_LINK_DIR, exist_ok=True)

    meals = ["Breakfast", "Lunch", "Dinner"]

    with pd.ExcelWriter(DISPATCH_LINK_FILE, engine="openpyxl", mode="w") as writer:
        for meal in meals:
            meal_df = orders_df[
                orders_df["Meal_Type"].astype(str).str.strip().str.title() == meal
            ].copy()

            if meal_df.empty:
                # Write empty sheet so all 3 tabs always exist
                pd.DataFrame(columns=[
                    "Customer Name", "Phone", "Full Address",
                    "Area", "Landmark", "Maps Link", "Special Notes"
                ]).to_excel(writer, index=False, sheet_name=meal)
                continue

            # Standardise
            meal_df["Customer_Name"] = (
                meal_df["Customer_Name"].astype(str).str.strip().str.title()
            )

            dispatch = meal_df[[
                "Customer_Name", "Phone", "Full_Address",
                "Area", "Landmark", "Maps_Link", "Special_Notes"
            ]].copy()

            dispatch.columns = [
                "Customer Name", "Phone", "Full Address",
                "Area", "Landmark", "Maps Link", "Special Notes"
            ]

            # Flag missing critical info
            for col in ["Phone", "Full Address"]:
                dispatch[col] = dispatch[col].replace({"": "⚠️ MISSING", "nan": "⚠️ MISSING"})

            dispatch.to_excel(writer, index=False, sheet_name=meal)

    _format_dispatch(DISPATCH_LINK_FILE, date_str)
    print(f"✅ Dispatch file updated → {DISPATCH_LINK_FILE}")


def _format_dispatch(file_path: str, date_str: str):
    wb   = load_workbook(file_path)
    thin = Side(border_style="thin", color="000000")
    red  = Font(color="FF0000", bold=True)

    for ws in wb.worksheets:
        ws.insert_rows(1)
        ws["A1"] = f"DISPATCH LIST — {date_str}"
        ws["A1"].font = Font(bold=True, size=14, color="FF0000")

        for cell in ws[2]:
            cell.font = Font(bold=True)

        for row in ws.iter_rows(min_row=3):
            for cell in row:
                if "⚠️" in str(cell.value):
                    cell.font = red
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) for c in col if c.value), default=14
            )
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

    wb.save(file_path)
