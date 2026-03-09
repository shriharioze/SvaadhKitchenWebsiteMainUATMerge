# payment_summary.py  (New System)
# Reads SK_Orders from Google Sheet.
# For a given date range (or auto-detected 10-day period), produces:
#   1. payment_summary_YYYY-MM-DD_to_YYYY-MM-DD.xlsx  — per-customer totals
#   2. Master_All_Years.xlsx                           — running invoice ledger
#   3. Monthly consolidated summary
#
# Works standalone (python payment_summary.py) or imported in Colab.

import os
import re
import json
import hashlib
import gspread
import pandas as pd
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from google.oauth2.service_account import Credentials

# ── CONFIG ───────────────────────────────────────────────────
SERVICE_ACCOUNT_FILE = "service_account.json"
SHEET_ID             = "17X7JOrMe1Oj_QykH7mk6UGoBjuguLfyC1RmKYATDXlI"
TAB_ORDERS           = "SK_Orders"

BASE_PATH    = os.path.join(os.getcwd(), "Processed_Orders", "Payment_Summary")
MASTER_FILE  = os.path.join(BASE_PATH, "Master_All_Years.xlsx")
PROC_LOG     = os.path.join(BASE_PATH, "processed_files.json")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# ── AUTH ─────────────────────────────────────────────────────
def _auth():
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return gspread.authorize(creds)

# ── PERIOD HELPERS ───────────────────────────────────────────
def current_period_dates() -> tuple:
    """Return (start_date, end_date) for today's 10-day billing period."""
    today = date.today()
    y, m, d = today.year, today.month, today.day
    if d <= 10:
        return date(y, m, 1), date(y, m, 10)
    elif d <= 20:
        return date(y, m, 11), date(y, m, 20)
    else:
        import calendar
        last = calendar.monthrange(y, m)[1]
        return date(y, m, 21), date(y, m, last)

def period_label(start: date, end: date) -> str:
    return f"{start.strftime('%Y-%m-%d')}_to_{end.strftime('%Y-%m-%d')}"

# ── FETCH ORDERS ─────────────────────────────────────────────
def fetch_orders_for_period(start: date, end: date) -> pd.DataFrame:
    gc  = _auth()
    sh  = gc.open_by_key(SHEET_ID)
    ws  = sh.worksheet(TAB_ORDERS)
    df  = pd.DataFrame(ws.get_all_records())

    if df.empty:
        return df

    df.columns = df.columns.str.strip()
    df["Order_Date"] = pd.to_datetime(df["Order_Date"], errors="coerce").dt.date
    df = df[
        (df["Order_Date"] >= start) & (df["Order_Date"] <= end)
    ].copy()
    return df

# ── AGGREGATE PER CUSTOMER ───────────────────────────────────
def aggregate_customer_totals(df: pd.DataFrame) -> pd.DataFrame:
    """One row per customer: total billed, pending, paid, payment freq."""
    if df.empty:
        return pd.DataFrame()

    df["Net_Total"]   = pd.to_numeric(df["Net_Total"],   errors="coerce").fillna(0)

    grp = df.groupby(["Customer_Name", "Phone", "Payment_Freq"]).agg(
        Total_Amount   = ("Net_Total",        "sum"),
        Pending_Amount = ("Net_Total",         lambda s: s[
            df.loc[s.index, "Payment_Status"].astype(str).str.lower() != "paid"
        ].sum()),
    ).reset_index()

    grp["Total_Amount"]   = grp["Total_Amount"].round(2)
    grp["Pending_Amount"] = grp["Pending_Amount"].round(2)
    grp["Payment_Status"] = grp["Pending_Amount"].apply(
        lambda x: "Paid" if x == 0 else "Pending"
    )
    return grp

# ── MASTER FILE HELPERS ───────────────────────────────────────
def _load_master() -> pd.DataFrame:
    if os.path.exists(MASTER_FILE):
        df = pd.read_excel(MASTER_FILE, dtype=str)
        df["Amount"]  = df["Amount"].astype(float)
        df["Pending"] = df["Pending"].astype(float)
        if "Payment Freq" not in df.columns:
            df["Payment Freq"] = ""
        return df
    return pd.DataFrame(columns=[
        "Invoice ID", "Hash Key", "Customer Name", "Phone",
        "Date", "Amount", "Status", "Pending", "Payment Freq"
    ])

def _save_master(df: pd.DataFrame):
    os.makedirs(os.path.dirname(MASTER_FILE), exist_ok=True)
    df.to_excel(MASTER_FILE, index=False)
    _auto_col_width(MASTER_FILE)

def _hash(customer: str, phone: str, date_str: str) -> str:
    """Unique key per customer+phone+period-end-date — amount excluded so corrections don't create duplicates."""
    key = f"{customer.lower().strip()}_{phone}_{date_str}"
    return hashlib.md5(key.encode()).hexdigest()

def _next_invoice_id(master_df: pd.DataFrame) -> str:
    """Always finds the true maximum invoice number — safe against row deletions and sorting."""
    year = datetime.now().year
    if master_df.empty:
        return f"INV-{year}-0001"
    try:
        nums = (
            master_df["Invoice ID"]
            .dropna()
            .astype(str)
            .str.extract(r"INV-\d{4}-(\d+)")[0]
            .dropna()
            .astype(int)
        )
        last_num = int(nums.max()) if not nums.empty else 0
    except Exception:
        last_num = 0
    return f"INV-{year}-{last_num + 1:04d}"

# ── MAIN PROCESSOR ───────────────────────────────────────────
def process_period(
    start: date = None,
    end:   date = None,
    confirm: bool = True,
) -> tuple:
    """
    Process a billing period.
    Returns (master_df, summary_df, output_path).
    """
    if start is None or end is None:
        start, end = current_period_dates()

    label     = period_label(start, end)
    end_str   = end.strftime("%Y-%m-%d")
    year      = end.year
    month_name= end.strftime("%B")

    print(f"\n📅 Period: {start} → {end}")
    print(f"📘 Master: {MASTER_FILE}")

    if confirm:
        ans = input("⚠️  Proceed with this period? (yes/no): ").strip().lower()
        if ans != "yes":
            print("❌ Cancelled.")
            return None, None, None

    # Fetch
    df = fetch_orders_for_period(start, end)
    if df.empty:
        print(f"ℹ️  No orders found for this period.")
        return None, None, None

    agg = aggregate_customer_totals(df)
    if agg.empty:
        print("⚠️  Aggregation returned empty.")
        return None, None, None

    # Update master
    master_df = _load_master()

    for _, row in agg.iterrows():
        customer = str(row["Customer_Name"]).strip()
        phone    = str(row["Phone"]).strip()
        amount   = float(row["Total_Amount"])
        status   = str(row["Payment_Status"])
        pending  = float(row["Pending_Amount"])
        freq     = str(row.get("Payment_Freq", ""))
        hk       = _hash(customer, phone, end_str)

        existing = master_df[master_df["Hash Key"] == hk]
        if not existing.empty:
            idx = existing.index[0]
            if master_df.at[idx, "Status"] == "Pending" and status == "Paid":
                master_df.at[idx, "Status"]  = "Paid"
                master_df.at[idx, "Pending"] = 0.0
            continue

        inv_id = _next_invoice_id(master_df)
        master_df = pd.concat([master_df, pd.DataFrame([{
            "Invoice ID":   inv_id,
            "Hash Key":     hk,
            "Customer Name":customer,
            "Phone":        phone,
            "Date":         end_str,
            "Amount":       amount,
            "Status":       status,
            "Pending":      pending,
            "Payment Freq": freq,
        }])], ignore_index=True)

    _save_master(master_df)
    print(f"✅ Master updated → {MASTER_FILE}")

    # Per-period summary file
    month_dir   = os.path.join(BASE_PATH, str(year), month_name)
    os.makedirs(month_dir, exist_ok=True)
    summary_path= os.path.join(month_dir, f"payment_summary_{label}.xlsx")

    agg.to_excel(summary_path, index=False)
    _auto_col_width(summary_path)
    print(f"✅ Period summary → {summary_path}")

    # Monthly consolidated
    _update_monthly_consolidated(master_df, year, month_name)

    # Customer-level summary
    cust_summary = master_df.groupby("Customer Name").agg(
        Total_Invoices = ("Invoice ID", "count"),
        Total_Billed   = ("Amount",     "sum"),
        Total_Pending  = ("Pending",    "sum"),
    ).reset_index()
    cust_summary["Total_Paid"] = cust_summary["Total_Billed"] - cust_summary["Total_Pending"]

    print("\n--- PERIOD SUMMARY PREVIEW ---")
    print(cust_summary.tail(10).to_string(index=False))

    return master_df, cust_summary, summary_path

# ── MONTHLY CONSOLIDATED ─────────────────────────────────────
def _update_monthly_consolidated(master_df: pd.DataFrame, year: int, month_name: str):
    month_dir  = os.path.join(BASE_PATH, str(year), month_name, "Summary")
    os.makedirs(month_dir, exist_ok=True)
    out_file   = os.path.join(month_dir, f"payment_summary_consolidated_{month_name}_{year}.xlsx")

    monthly = master_df[
        master_df["Date"].astype(str).str.startswith(str(year))
    ].copy()
    monthly["Month"] = pd.to_datetime(monthly["Date"], errors="coerce").dt.strftime("%B")
    monthly = monthly[monthly["Month"] == month_name]

    consolidated = monthly.groupby("Customer Name").agg(
        Total_Invoices = ("Invoice ID", "count"),
        Total_Billed   = ("Amount",     "sum"),
        Total_Pending  = ("Pending",    "sum"),
    ).reset_index()
    consolidated["Total_Paid"] = consolidated["Total_Billed"] - consolidated["Total_Pending"]

    consolidated.to_excel(out_file, index=False)
    _auto_col_width(out_file)
    print(f"✅ Monthly consolidated → {out_file}")

# ── FORMATTING ───────────────────────────────────────────────
def _auto_col_width(path: str):
    try:
        wb = load_workbook(path)
        ws = wb.active
        thin = Side(border_style="thin", color="000000")
        for idx, col in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 3
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(path)
    except Exception as e:
        print(f"⚠️  Formatting skipped: {e}")

# ── STANDALONE ───────────────────────────────────────────────
if __name__ == "__main__":
    process_period(confirm=True)
